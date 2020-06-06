# -*- coding: utf-8 -*-
import sys
from result_storage import result_storage
import datetime
import tkinter as tk
from tkinter import *
from pymouse import PyMouse
from pykeyboard import PyKeyboard
import time
import re
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference
import logging

result_list = result_storage()
filename = 'logXX.log'
#测试节点
shortID_dict = []
#心跳节点
shortID_dict_no_use = []

def route_save():
    global  filename
    filename = e.get()
    excel_create()
    text.insert((1.0), '路径确认 ：%s \n'%(filename))
    text.insert((1.0), '---------------------------------------\n')
    window.update()
def node_laoding():
    file1 = '节点列表.txt'
    global shortID_dict,shortID_dict_no_use
    node_type = 0 #1 表示测试节点，2  表示心跳节点
    f = open(file1, encoding='utf-8')
    rows = len(f.readlines())
    f = open(file1, encoding='utf-8')
    rows_cnt = 0  # 读取行数计数
    text.insert((1.0), '---------------------------------------\n')
    while rows_cnt < rows:
        str_check = f.readline()
        if str_check[:4] == '测试节点':
            node_type = 1
        if str_check[:4] == '心跳节点':
            node_type = 2
        if node_type == 1 and str_check[:4] != '测试节点':
            shortID_dict.append(str_check[:4])
        elif node_type == 2 and str_check[:4] != '心跳节点':
            shortID_dict_no_use.append(str_check[:4])
        rows_cnt = rows_cnt+1
    text.insert((1.0), '%s\n'%(shortID_dict_no_use))
    text.insert((1.0), '心跳节点：\n')
    text.insert((1.0), '%s\n'%(shortID_dict))
    text.insert((1.0), '待测节点：\n')
    text.insert((1.0), '---------------------------------------\n')
    window.update()
def excel_create():
    global FILE_PATH
    global wb
    FILE_NAME = 'Aiweb Test Result'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Ping Test Result'
    #ws1 = wb.create_sheet(title=u'Ping Test Result')
    ws2 = wb.create_sheet(title=u'20s Test Result')
    ws3 = wb.create_sheet(title=u'120s Test Result')
    ws4 = wb.create_sheet(title=u'120s Test Result Backup')
    tmp_dict = {'节点编号': 'A1', '测试时间和成功率': 'B1',}
    for key, des in tmp_dict.items():
        ws1[des] = key
    for key, des in tmp_dict.items():
        ws2[des] = key
    for key, des in tmp_dict.items():
        ws3[des] = key
    for key, des in tmp_dict.items():
        ws4[des] = key
    FILE_TIME = datetime.datetime.now()
    # 将时间转换成字符串并修改格式
    FILE_TIME = repr(FILE_TIME).replace('datetime.datetime', '')
    FILE_TIME = FILE_TIME.replace(', ', '-')
    # 生成文件名
    FILE_PATH = FILE_NAME + FILE_TIME + ".xlsx"
    wb.save(FILE_PATH)
    pass
def excel_write(sheet_num = 0,row = 1,col = 1,data = ''):
    column = 'A'
    sheet = wb.worksheets[sheet_num]
    column = get_char(col-1)
    sheet[column+str(row)]= data
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet[column+str(row)].alignment = alignment
    #wb.save(FILE_PATH)
def get_char(number):
    factor, moder = divmod(number, 26) # 26 字母个数
    mod_char = chr(moder + 65)          # 65 -> 'A'
    if factor != 0:
        mod_char = get_char(factor-1) + mod_char # factor - 1 : 商为有效值时起始数为 1 而余数是 0
    return mod_char
def set_cell_color(column = 'a',row = 1,data = 100,test_num = 0):
    sheet = wb.worksheets[test_num]
    color = ['F8696B','F98370','FA9D75','FCB77A','FDD17F','FFEB84','C0E383','C1DA81','A2D07F','83C77D','63BE7B','F8696B']  #RGB颜色编码，0%或大于100%为红色
    data = int(data*10)
    if data >10:
        data = 11
    fill = PatternFill(start_color=color[data], end_color=color[data], fill_type='solid')  # 填充黄色
    sheet[column+str(row)].fill = fill
def result_to_excel(target_node = '',data_time = '',success_rate = 100.00,test_cycle_cnt = 0,test_num = 0,shortID_dict1 = [],backup = 0):
    column = 'A'
    sheet = wb.worksheets[test_num]
    column = get_char(test_cycle_cnt)
    sheet.column_dimensions[column].width = 12.0
    if backup == 0:
        if shortID_dict1.index(target_node) == 0:
            sheet.row_dimensions[shortID_dict1.index(target_node) + 2].height = 25
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 2,
                        col=test_cycle_cnt + 1,
                        data=data_time)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=1,
                        data=target_node)
            sheet[column+str(shortID_dict1.index(target_node) + 3)].number_format = '00.00%'
            set_cell_color(column, shortID_dict1.index(target_node) + 3, success_rate,test_num)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=test_cycle_cnt + 1,
                        data=success_rate)
        else:
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=1,
                        data=target_node)
            sheet[column + str(shortID_dict1.index(target_node) + 3)].number_format = '00.00%'
            set_cell_color(column, shortID_dict1.index(target_node) + 3, success_rate,test_num)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=test_cycle_cnt + 1,
                        data=success_rate)
    elif backup == 1:
        sheet.row_dimensions[(shortID_dict1.index(target_node) + 1)*2].height = 25
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1)*2,
                    col=test_cycle_cnt + 1,
                    data=data_time)
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1)*2+1,
                    col=1,
                    data=target_node)
        sheet[column + str((shortID_dict1.index(target_node) + 1)*2+1)].number_format = '00.00%'
        set_cell_color(column, (shortID_dict1.index(target_node) + 1)*2+1, success_rate,test_num)
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1)*2+1,
                    col=test_cycle_cnt + 1,
                    data=success_rate)
def creat_bar_chart(test_num = 0,chart_title = ''):
    sheet = wb.worksheets[test_num]
    row_max = sheet.max_row  # 获取行数
    col_max = sheet.max_column  # 获取列数
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = chart_title
    chart1.y_axis.title = '成功率'
    chart1.x_axis.title = '节点编号'
    data = Reference(sheet, min_col=2, min_row=2, max_row=row_max, max_col=col_max)
    cats = Reference(sheet, min_col=1, min_row=3, max_row=row_max)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet.add_chart(chart1,'A17')

def downlink_test(shortID_dict,cmd,packet_cnt_range = 3,packet_intervals = 1):
    m = PyMouse()
    k = PyKeyboard()
    global t
    shortID_cnt = 0  # 短地址键值计数
    packet_cnt = 1  # 包计数
    send_packet = 'a'  # 组包用字符串
   # m.click(132, 29, 1, 1)  # 激活窗口，不知道还有啥好方法
    while packet_cnt <= packet_cnt_range:
        for shortID_cnt in range(len(shortID_dict)):
            m.click(370, 850, 1, 1)  # 激活窗口，不知道还有啥好方法
            packet_cnt_str = str(packet_cnt)
            shortID_dict[shortID_cnt] = shortID_dict[shortID_cnt].zfill(4)
            if cmd == 'ping':
                send_packet = 'at+ping,' + shortID_dict[shortID_cnt] + ',08' + shortID_dict[shortID_cnt] + '38' + packet_cnt_str.zfill(8)
            elif cmd == 'st1':  # 20s上行测试开启
                send_packet = 'at+st,' + shortID_dict[shortID_cnt] + ',Wd40001'
            elif cmd == 'st2':  # 120s上行测试开启
                send_packet = 'at+st,' + shortID_dict[shortID_cnt] + ',Wd50001'
            elif cmd == 'st3':  # 120s上行测试关闭
                send_packet = 'at+st,' + shortID_dict[shortID_cnt] + ',Wd50000'
            elif cmd == 'st4':  # 下行测试
                send_packet = 'at+st,' + shortID_dict[shortID_cnt] + ',Wd3' + packet_cnt_str.zfill(3) + '11111111111111111111111111'
            elif cmd == 'st5':  # 下行测试成功率回读
                send_packet = 'at+st,' + shortID_dict[shortID_cnt] + ',Wd6' + packet_cnt_str.zfill(3) + '1'
            text.insert((1.0), '·')
            window.update()
            k.type_string(send_packet)
            k.tap_key(k.enter_key)
            time.sleep(packet_intervals)
        packet_cnt = packet_cnt + 1
def downlink_sucess_rate(shortID_dict, dowmlink_cnt, filename='新建文本文档.txt', target_node='0001'):
    f = open(filename, encoding='utf-8')
    rows = len(f.readlines())
    f = open(filename, encoding='utf-8')
    sample_str = '     0000  75 33 30 30 30 31 31 31  31 31 31 31 31 31 31 31  u3000111 11111111'
    nun_per_nodes = 50.0  # 单节点测试次数
    rows_cnt = 0  # 读取行数计数
    valid_rows_cnt = 0  # 有效行计数
    node_num = 0
    data_time_now = '-'
    data_time = '-'
    packet_cnt = -1
    time_interval = 0
    dowmlink_cnt[target_node] = 0
    test_cycle_cnt = 0
    while rows_cnt < rows:
        str_check = f.readline().strip()
        rows_cnt = rows_cnt + 1
        if str_check[:5] == 'Frame':
            data_time_now = str_check[-26:-7]
        if str_check[:9] == 'recv_pong'and str_check[:13] !='recv_pong not':  # 定位目标数据的位置，当前为目标数据的上一行
            #node_num = hex(int(str_check[10:13]))  # 获取当前节点小名，并转换为16进制
            node_num = re.match(r'recv_pong (.*) ', str_check)
            try:
                node_num = hex(int(node_num.group(1)))
            except:
                print('节点格式转换异常')
                print()
                logging.error('节点格式转换异常:%s'%node_num.group(1))
            else:
                node_num = str(node_num).upper().replace('0X','').zfill(4)
                packet_cnt_now = str_check[-3:-1]  # 获取当前数据包计数
                if node_num == target_node:  # 比较当前数据包所属节点 Short ID 与 目标节点是否一致
                    if packet_cnt < int(packet_cnt_now):
                        packet_cnt = int(packet_cnt_now)
                        dowmlink_cnt[target_node] = dowmlink_cnt[target_node] + 1
                        data_time = data_time_now
                    else:
                        text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                              % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                 int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                        window.update()
                        print( '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                                    % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                       int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                        test_cycle_cnt +=1
                        success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
                        result_to_excel(target_node,data_time,success_rate,test_cycle_cnt,0,shortID_dict)
                        dowmlink_cnt[target_node] = 0
                        packet_cnt = 0
        if rows_cnt == rows - 1:
            text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                        % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                           int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
            window.update()
            print('节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                  % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                     int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
            test_cycle_cnt += 1
            success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
            result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 0, shortID_dict)
            dowmlink_cnt[target_node] = 0
            packet_cnt = 0
def uplink_sucess_rate(cmd, shortID_dict, dowmlink_cnt, filename='新建文本文档.txt', target_node='0001'):
    print("目标地址：%s" % (target_node))
    f = open(filename, encoding='utf-8')
    rows = len(f.readlines())
    f = open(filename, encoding='utf-8')
    sample_str = '     0000  75 33 30 30 30 31 31 31  31 31 31 31 31 31 31 31  u3000111 11111111'
    nun_per_nodes = 100.0  # 单节点测试次数
    rows_cnt = 0  # 读取行数计数
    valid_rows_cnt = 0  # 有效行计数
    data_time_now = '-'
    data_time = '-'
    packet_cnt = 0
    time_interval = 0
    test_cycle_cnt = 0
    while rows_cnt < rows:
        str_check = f.readline().strip()
        rows_cnt = rows_cnt + 1
        if str_check[:5] == 'Frame':
            data_time_now = str_check[-26:-7]
            print("-----------------------------------------------------------------------------")
            print("frame time:%s"%(data_time_now))
        if str_check == 'procBody application data':  # 定位目标数据的位置，当前为目标数据的上一行
            str_check = f.readline().strip()
            rows_cnt = rows_cnt + 1
            str_check = f.readline().strip()
            rows_cnt = rows_cnt + 1
            #定位数据
            # procBody application data
            # 2020-04-18 11:23:38.676549
            # 0000  75 31 30 36 30 31 31 31  31 31 31 31 31 31 31 31  u1060111 11111111
            print(str_check)
            print("-----------------------------------------------------------------------------")
            rows_cnt = rows_cnt + 1
            if len(str_check) == len(sample_str.strip()):  # 判断数据长度是否满足要求
                if str_check[-17:-15] == cmd:  # 判断是否为 20/120s 下行数据
                    print(str_check)
                    packet_cnt_now = str_check[-15:-12]  # 获取当前数据包计数
                    print("当前包计数：%s"%(packet_cnt_now))
                    while str_check[:5] != 'saddr':
                        str_check = f.readline().strip()
                        rows_cnt = rows_cnt + 1
                    print("节点编号：%s    目标节点编号：%s"%(str_check[6:10],target_node))
                    if str_check[6:10] == target_node:  # 比较当前数据包所属节点 Short ID 与 目标节点是否一致
                        if packet_cnt < int(packet_cnt_now):
                            packet_cnt = int(packet_cnt_now)
                            dowmlink_cnt[target_node] = dowmlink_cnt[target_node] + 1
                            data_time = data_time_now
                            print("计数：%d"%(dowmlink_cnt[target_node]))
                        else:
                            if cmd == 'u1':#20s间隔上行数据
                                if data_time == '-':
                                    print("time now:%s"%(data_time_now))
                                    data_time = data_time_now
                                time_interval = int(data_time_now[8:10])*24*3600+int(data_time_now[11:13])*3600+int(data_time_now[14:16])*60+int(data_time_now[17:19])\
                                                -int(data_time[8:10])*24*3600-int(data_time[11:13])*3600-int(data_time[14:16])*60-int(data_time[17:19])

                                if time_interval < 100:
                                    packet_cnt = int(packet_cnt_now)
                                    print('当前包计数',packet_cnt)
                                    dowmlink_cnt[target_node] = dowmlink_cnt[target_node] + 1
                                    data_time = data_time_now
                                else:
                                    # text.insert((1.0), 'time_interval%d\n' % (time_interval))
                                    text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                                          % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                             int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                                    window.update()
                                    test_cycle_cnt += 1
                                    success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
                                    result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 1,shortID_dict)
                                    dowmlink_cnt[target_node] = 0
                                    packet_cnt = 0
                            if cmd == 'u2':#120s间隔上行数据
                                if test_cycle_cnt == 0:
                                    test_cycle_cnt += 1
                                else:
                                    text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                                                % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                                   int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                                    success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
                                    result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 2,shortID_dict_no_use)
                                    result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 3,shortID_dict_no_use,1)
                                    test_cycle_cnt += 1
                                    window.update()
                                    dowmlink_cnt[target_node] = 0
                                    packet_cnt = 0
                        valid_rows_cnt = valid_rows_cnt + 1
        if rows_cnt == rows - 1:
            if cmd == 'u1':
                # text.insert((1.0), 'time_interval%d\n' % (time_interval))
                text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                            % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                               int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                window.update()
                test_cycle_cnt += 1
                success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
                result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 1, shortID_dict)
                dowmlink_cnt[target_node] = 0
                packet_cnt = 0
            if cmd == 'u2':#最后一组使用当前包计数作为改组的总包数
                print('end line，','当前包计数', packet_cnt)
                if packet_cnt == 0:
                    text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                                % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                   int(dowmlink_cnt[target_node]) / nun_per_nodes * 100, data_time))
                    success_rate = int(dowmlink_cnt[target_node]) / nun_per_nodes
                else:
                    text.insert((1.0), '节点编号：%s  成功次数:%s  成功率:%05.2f%% 测试时间：%s\n'
                                % (target_node, str(dowmlink_cnt[target_node]).rjust(2),
                                   int(dowmlink_cnt[target_node]) / packet_cnt * 100, data_time))
                    success_rate = int(dowmlink_cnt[target_node]) / packet_cnt
                if test_cycle_cnt ==0:
                    test_cycle_cnt += 1
                result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 2, shortID_dict_no_use)
                result_to_excel(target_node, data_time, success_rate, test_cycle_cnt, 3, shortID_dict_no_use, 1)
                window.update()
                dowmlink_cnt[target_node] = 0
                packet_cnt = 0


def task_info_show(task_info=''):
    text.insert((1.0), '%s  %s\n'%(task_info,datetime.datetime.now()))
    text.insert((1.0), '---------------------------------------\n')
    window.update()
def task_dowmlink_ping(interval=5):
    task_info_show('at+ping 自动发送测试开始')
    downlink_test(shortID_dict, 'ping', 50, interval)
    task_info_show('at+ping 自动发送测试结束')
def task_dowmlink_st(interval=5):
    task_info_show('下行测试开始')
    downlink_test(shortID_dict, 'st4', 50, interval)
    #downlink_test(shortID_dict, 'st5', 5, interval)
    task_info_show('下行测试结束')
def task_uplink_20s_on(interval=5):
    cnt = 0
    task_info_show('20s上行测试开始')
    downlink_test(shortID_dict, 'st1', 5, interval)
    while cnt != 200:
        cnt = cnt+1
        time.sleep(10)  # 阻塞式休眠
        task_info_show('20s上行测试中。。。')
    task_info_show('20s上行测试结束')
def task_uplink_120s_on_1(interval=5):
    task_info_show('待测节点上行心跳开启中')
    downlink_test(shortID_dict, 'st2', 5, 5)
    task_info_show('待测节点上行心跳开启完成')
def task_uplink_120s_on_2(interval=5):
    task_info_show('非待测节点上行心跳开启中')
    downlink_test(shortID_dict_no_use, 'st2', 5, interval)
    task_info_show('非待测节点上行心跳开启完成')
def task_uplink_120s_off(interval=5):
    task_info_show('待测节点上行心跳关闭中')
    downlink_test(shortID_dict, 'st3', 5, interval)
    task_info_show('待测节点上行心跳关闭完成')
def downlink_sucess_rate_ping():
    time_write_flag = 1
    for shortID_cnt in range(len(shortID_dict)):
        target_node = shortID_dict[shortID_cnt]
        downlink_sucess_rate(shortID_dict, result_list, filename, target_node)
        text.insert((1.0), '---------------------------------------\n')
    creat_bar_chart(0,'Ping 成功率统计图')
    wb.save(FILE_PATH)
    task_info_show('下行测试统计结果')
def uplink_sucess_rate_20s():
    for shortID_cnt in range(len(shortID_dict)):
        target_node = shortID_dict[shortID_cnt]
        uplink_sucess_rate('u1', shortID_dict, result_list, filename, target_node)
        text.insert((1.0), '---------------------------------------\n')
    creat_bar_chart(1,'20s 上行成功率统计图')
    wb.save(FILE_PATH)
    task_info_show('20s上行测试统计结果')
def uplink_sucess_rate_120s():
    for shortID_cnt in range(len(shortID_dict_no_use)):
        target_node = shortID_dict_no_use[shortID_cnt]
        uplink_sucess_rate('u2', shortID_dict, result_list, filename, target_node)
        text.insert((1.0), '---------------------------------------\n')
    creat_bar_chart(2,'120s 上行成功率统计图')
    wb.save(FILE_PATH)
    task_info_show('120s上行测试统计结果')
def all_sucess_rate():
    task_info_show('一键统计开始')
    #downlink_sucess_rate_ping()
    uplink_sucess_rate_20s()
    uplink_sucess_rate_120s()
    task_info_show('一键统计结束')
def comprehensive_test():
    task_info_show('综合测试开始')
    task_uplink_120s_off()
    task_uplink_120s_on_2()
    task_dowmlink_ping()
    task_uplink_20s_on()
    task_dowmlink_st()
    task_dowmlink_ping()
    task_uplink_20s_on()
    task_dowmlink_st()
    task_uplink_120s_on_1()
    task_info_show('综合测试结束')

# 用户界面
window = tk.Tk()
window.title('Aiweb 测试程序 V2.0')
window.geometry('700x500')  # 这里的乘是小x
l1 = tk.Label(window,
             font=('Arial', 12),
             width=20,
             height=2,
             text='Aiweb 测试程序 V2.0').place(relx=0.5, rely=0.05, anchor=CENTER)
l2 = tk.Label(window,
             fg='black',
             font=('Arial', 12),
             width=22,
             height=2,
             text='待统计文件路径：').place(relx=0.12, rely=0.12, anchor=CENTER)
e = tk.Entry(window,
             width=60,
             show=None)
e.place(relx=0.52, rely=0.12, anchor=CENTER)
b1 = tk.Button(window,
               text='综合测试',
               width=15,
               height=2,
               command=comprehensive_test).place(relx=0.1, rely=0.22, anchor=CENTER)
b2 = tk.Button(window,
               text='ping成功率统计',
               width=15,
               height=2,
               command=downlink_sucess_rate_ping).place(relx=0.1, rely=0.35, anchor=CENTER)
b3 = tk.Button(window,
               text='20s上行成功率统计',
               width=15,
               height=2,
               command=uplink_sucess_rate_20s).place(relx=0.1, rely=0.48, anchor=CENTER)
b4 = tk.Button(window,
               text='120s上行成功率统计',
               width=15,
               height=2,
               command=uplink_sucess_rate_120s).place(relx=0.1, rely=0.61, anchor=CENTER)
b5 = tk.Button(window,
               text='一键统计',
               width=15,
               height=2,
               command=all_sucess_rate).place(relx=0.1, rely=0.74, anchor=CENTER)
b6 = tk.Button(window,
               text='确认输入',
               width=10,
               height=1,
               command=route_save).place(relx=0.9, rely=0.12, anchor=CENTER)
text = tk.Text(window,
            width=75,
            height=25)
text.place(relx=0.60, rely=0.5, anchor=CENTER)
node_laoding()
text.insert((1.0), '说明：\n1.综合测试测试项包含ping测试、20s上行测试和下行测试，轮询进行两组。\n2.log文件与测试程序处于相同路径时，可以只输入带后缀的文件名，否则需要输入完整路径或相对路径。\n3.成功率统计结果会自动生成excel表格，与测试工具保存在同一路径。\n')
text.insert((1.0), 'Aiweb 测试程序 2020-01-08 Version 2.1\n')
text.insert((1.0), '---------------------------------------\n')
window.mainloop()