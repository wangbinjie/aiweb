# -*- coding: utf-8 -*-
import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
import json
import GlobalPara as G_Para


def node_laoding():
    global shortID_dict, shortID_hertBeat
    result_dict = {}
    file2 = 'nodelist.json'
    try:
        with open(file2, 'r', encoding="utf-8") as file:
            nodelist = json.load(file)
    except BaseException:
        log_list('节点载入失败')
    shortID_dict = nodelist.get('测试节点')
    shortID_hertBeat = nodelist.get('心跳节点')
    G_Para.set_value('shortID', shortID_dict)
    G_Para.set_value('shortID_hertBeat', shortID_hertBeat)
    shortID = G_Para.get_value('shortID')
    shortID_hertBeat = G_Para.get_value('shortID_hertBeat')
    task_info_show('载入节点列表')
    log_list('待测节点：')
    log_list('%s' % shortID)
    log_list('心跳节点：')
    log_list('%s' % shortID_hertBeat)
    log_list('---------------------------------------')
    for key in shortID:
        result_dict.setdefault(key, 0)
    for key in shortID_hertBeat:
        result_dict.setdefault(key, 0)
    G_Para.set_value('result_dict', result_dict)


def readme_loading():
    with open('readme.txt', 'r',encoding='utf-8') as f:
        readme = f.read()
        log_list(readme)

def excel_create():
    global FILE_PATH
    global wb
    FILE_PATH = r'.\TestResults\\'
    FILE_NAME = 'Aiweb Test Result'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Ping Test Result'
    # ws1 = wb.create_sheet(title=u'Ping Test Result')
    ws2 = wb.create_sheet(title=u'20s Test Result')
    ws3 = wb.create_sheet(title=u'120s Test Result')
    ws4 = wb.create_sheet(title=u'120s Test Result Backup')
    tmp_dict = {'节点编号': 'A1', '测试时间和成功率': 'B1'}
    for key, des in tmp_dict.items():
        ws1[des] = key
    for key, des in tmp_dict.items():
        ws2[des] = key
    for key, des in tmp_dict.items():
        ws3[des] = key
    for key, des in tmp_dict.items():
        ws4[des] = key
    FILE_TIME = datetime.datetime.now()
    # 将时间转换成字符串并修改格式
    FILE_TIME = repr(FILE_TIME).replace('datetime.datetime', '')
    FILE_TIME = FILE_TIME.replace(', ', '-')
    # 生成文件名
    FILE_PATH = FILE_PATH + FILE_NAME + FILE_TIME + ".xlsx"
    wb.save(FILE_PATH)
    task_info_show('Excel 创建成功')


def excel_write(sheet_num=0, row=1, col=1, data=''):
    column = 'A'
    sheet = wb.worksheets[sheet_num]
    column = get_char(col - 1)
    sheet[column + str(row)] = data
    alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True)
    sheet[column + str(row)].alignment = alignment
    # wb.save(FILE_PATH)


def excel_save():
    wb.save(FILE_PATH)


def get_char(number):
    factor, moder = divmod(number, 26)  # 26 字母个数
    mod_char = chr(moder + 65)  # 65 -> 'A'
    if factor != 0:
        # factor - 1 : 商为有效值时起始数为 1 而余数是 0
        mod_char = get_char(factor - 1) + mod_char
    return mod_char


def set_cell_color(column='a', row=1, data=100, test_num=0):
    sheet = wb.worksheets[test_num]
    color = [
        'F8696B',
        'F98370',
        'FA9D75',
        'FCB77A',
        'FDD17F',
        'FFEB84',
        'C0E383',
        'C1DA81',
        'A2D07F',
        '83C77D',
        '63BE7B',
        'F8696B']  # RGB颜色编码，0%或大于100%为红色
    data = int(data * 10)
    if data > 10:
        data = 11
    fill = PatternFill(
        start_color=color[data],
        end_color=color[data],
        fill_type='solid')  # 填充黄色
    sheet[column + str(row)].fill = fill


def result_to_excel(
        target_node='',
        data_time='',
        success_rate=100.00,
        test_cycle_cnt=0,
        test_num=0,
        shortID_dict1=[],
        backup=0):
    column = 'A'
    sheet = wb.worksheets[test_num]
    column = get_char(test_cycle_cnt)
    sheet.column_dimensions[column].width = 12.0
    if backup == 0:
        if shortID_dict1.index(target_node) == 0:
            sheet.row_dimensions[shortID_dict1.index(
                target_node) + 2].height = 25
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 2,
                        col=test_cycle_cnt + 1,
                        data=data_time)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=1,
                        data=target_node)
            sheet[column +
                  str(shortID_dict1.index(target_node) +
                      3)].number_format = '00.00%'
            set_cell_color(column, shortID_dict1.index(
                target_node) + 3, success_rate, test_num)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=test_cycle_cnt + 1,
                        data=success_rate)
        else:
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=1,
                        data=target_node)
            sheet[column +
                  str(shortID_dict1.index(target_node) +
                      3)].number_format = '00.00%'
            set_cell_color(column, shortID_dict1.index(
                target_node) + 3, success_rate, test_num)
            excel_write(sheet_num=test_num,
                        row=shortID_dict1.index(target_node) + 3,
                        col=test_cycle_cnt + 1,
                        data=success_rate)
    elif backup == 1:
        sheet.row_dimensions[(shortID_dict1.index(
            target_node) + 1) * 2].height = 25
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1) * 2,
                    col=test_cycle_cnt + 1,
                    data=data_time)
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1) * 2 + 1,
                    col=1,
                    data=target_node)
        sheet[column + str((shortID_dict1.index(target_node) + 1)
                           * 2 + 1)].number_format = '00.00%'
        set_cell_color(
            column,
            (shortID_dict1.index(target_node) + 1) * 2 + 1,
            success_rate,
            test_num)
        excel_write(sheet_num=test_num,
                    row=(shortID_dict1.index(target_node) + 1) * 2 + 1,
                    col=test_cycle_cnt + 1,
                    data=success_rate)


def creat_bar_chart(test_num=0, chart_title=''):
    sheet = wb.worksheets[test_num]
    row_max = sheet.max_row  # 获取行数
    col_max = sheet.max_column  # 获取列数
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = chart_title
    chart1.y_axis.title = '成功率'
    chart1.x_axis.title = '节点编号'
    data = Reference(
        sheet,
        min_col=2,
        min_row=2,
        max_row=row_max,
        max_col=col_max)
    cats = Reference(sheet, min_col=1, min_row=3, max_row=row_max)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet.add_chart(chart1, 'A17')


def log_list(log):
    if G_Para.get_value('mylog') is None:
        G_Para.set_value('mylog', [])
    mylog = G_Para.get_value('mylog')
    mylog.append(log)
    G_Para.set_value('mylog', mylog)


def task_info_show(task_info=''):
    task_info = task_info + '   ' + str(datetime.datetime.now())
    log_list('%s' % (task_info))
    log_list('-' * len(task_info))
